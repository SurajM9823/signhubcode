from django.shortcuts import render, HttpResponse
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from clients.models import Board, Client, FiscalYear, BoardType, Brand
from django.core.exceptions import ValidationError
import openpyxl
from openpyxl.utils import get_column_letter
from django.db import transaction
from decimal import Decimal
import io
import logging

# Set up logging
logger = logging.getLogger(__name__)

def excel_upload(request):
    """
    Renders the Excel upload page.
    """
    return render(request, 'excel_upload.html')

@csrf_exempt
def download_excel_template(request):
    """
    Generates and returns an Excel template for board data upload, optionally tailored to a specific client.
    """
    try:
        fiscal_year = FiscalYear.objects.filter(is_active=True).first()
        if not fiscal_year:
            return JsonResponse({'success': False, 'message': 'No active fiscal year found'}, status=400)

        # Get client_id from query parameters (optional)
        client_id = request.GET.get('client_id')
        
        # Fetch board types and brands for the active fiscal year
        board_types = BoardType.objects.filter(fiscal_year=fiscal_year).values_list('name', flat=True)
        if client_id:
            try:
                client = Client.objects.get(id=client_id, fiscal_year=fiscal_year)
                clients = [(client.id, client.name)]
                brands = Brand.objects.filter(fiscal_year=fiscal_year, client=client).values_list('name', flat=True)
            except Client.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'Client ID {client_id} not found'}, status=404)
        else:
            clients = Client.objects.filter(fiscal_year=fiscal_year).values_list('id', 'name')
            brands = Brand.objects.filter(fiscal_year=fiscal_year).values_list('name', flat=True)

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Board Data Template"

        # Define headers
        headers = [
            'Client ID', 'Address', 'Shop Name', 'Brand Name', 'Board Type Name',
            'Length (ft)', 'Breadth (ft)', 'Quantity', 'Rate (NPR/sqft)', 'Discount (NPR)'
        ]
        for col, header in enumerate(headers, start=1):
            ws[f'{get_column_letter(col)}1'] = header

        # Add sample data
        ws['A2'] = clients[0][0] if clients else ''
        ws['B2'] = '123 Main Street, Kathmandu'
        ws['C2'] = 'City Electronics'
        ws['D2'] = brands[0] if brands else ''
        ws['E2'] = board_types[0] if board_types else ''
        ws['F2'] = 12.5
        ws['G2'] = 6.0
        ws['H2'] = 2
        ws['I2'] = 150.0
        ws['J2'] = 50.0

        # Add data validation for Client ID
        from openpyxl.worksheet.datavalidation import DataValidation
        client_ids = ','.join(str(client[0]) for client in clients)
        dv_client = DataValidation(type="list", formula1=f'"{client_ids}"', allow_blank=True)
        dv_client.add('A2:A1000')
        ws.add_data_validation(dv_client)

        # Add data validation for Board Type
        board_types_list = ','.join(f'"{bt}"' for bt in board_types)
        dv_board_type = DataValidation(type="list", formula1=f'"{board_types_list}"', allow_blank=True)
        dv_board_type.add('E2:E1000')
        ws.add_data_validation(dv_board_type)

        # Add data validation for Brand
        brands_list = ','.join(f'"{b}"' for b in brands)
        dv_brand = DataValidation(type="list", formula1=f'"{brands_list}"', allow_blank=True)
        dv_brand.add('D2:D1000')
        ws.add_data_validation(dv_brand)

        # Save workbook to a byte stream
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Prepare response
        response = HttpResponse(
            content=output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=board_data_template.xlsx'
        return response
    except Exception as e:
        logger.error(f"Error generating Excel template: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error generating template: {str(e)}'}, status=500)

@csrf_exempt
def upload_excel(request):
    """
    Handles the upload and processing of Excel files for bulk board data import.
    Validates that Client ID in Excel matches the selected client from the UI.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

    try:
        fiscal_year = FiscalYear.objects.filter(is_active=True).first()
        if not fiscal_year:
            return JsonResponse({'success': False, 'message': 'No active fiscal year found'}, status=400)

        # Get selected client_id from form data
        selected_client_id = request.POST.get('client_id')
        if not selected_client_id:
            return JsonResponse({'success': False, 'message': 'No client selected'}, status=400)

        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'message': 'No file uploaded'}, status=400)

        excel_file = request.FILES['file']
        if not excel_file.name.endswith('.xlsx'):
            return JsonResponse({'success': False, 'message': 'Invalid file format. Please upload an .xlsx file'}, status=400)

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # Expected headers
        expected_headers = [
            'Client ID', 'Address', 'Shop Name', 'Brand Name', 'Board Type Name',
            'Length (ft)', 'Breadth (ft)', 'Quantity', 'Rate (NPR/sqft)', 'Discount (NPR)'
        ]
        headers = [cell.value for cell in ws[1]]
        if headers != expected_headers:
            return JsonResponse({'success': False, 'message': 'Invalid Excel template. Please use the provided template'}, status=400)

        errors = []
        boards_to_save = []

        # Cache database queries
        clients = {str(c.id): c for c in Client.objects.filter(fiscal_year=fiscal_year)}
        brands = {b.name: b for b in Brand.objects.filter(fiscal_year=fiscal_year)}
        board_types = {bt.name: bt for bt in BoardType.objects.filter(fiscal_year=fiscal_year)}

        # Validate selected client exists
        if selected_client_id not in clients:
            return JsonResponse({'success': False, 'message': f'Invalid selected Client ID {selected_client_id}'}, status=400)

        with transaction.atomic():
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(cell is None for cell in row):
                    continue  # Skip empty rows

                client_id, address, shop_name, brand_name, board_type_name, length, breadth, quantity, rate, discount = row

                # Validate required fields
                if not all([client_id, address, shop_name, brand_name, board_type_name, length, breadth, quantity, rate]):
                    errors.append(f"Row {row_idx}: Missing required fields")
                    continue

                # Validate Client ID matches selected client
                client_id_str = str(client_id)
                if client_id_str != selected_client_id:
                    errors.append(f"Row {row_idx}: Client ID {client_id} does not match selected client ID {selected_client_id}")
                    continue

                # Validate foreign keys
                if client_id_str not in clients:
                    errors.append(f"Row {row_idx}: Invalid Client ID {client_id}")
                    continue

                if brand_name not in brands:
                    errors.append(f"Row {row_idx}: Invalid Brand Name {brand_name}")
                    continue

                if board_type_name not in board_types:
                    errors.append(f"Row {row_idx}: Invalid Board Type Name {board_type_name}")
                    continue

                # Validate numeric fields
                try:
                    length = Decimal(str(length))
                    breadth = Decimal(str(breadth))
                    quantity = int(quantity)
                    rate = Decimal(str(rate))
                    discount = Decimal(str(discount or 0))
                    if length <= 0 or breadth <= 0 or quantity <= 0 or rate <= 0:
                        errors.append(f"Row {row_idx}: Numeric fields must be positive")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Row {row_idx}: Invalid numeric values")
                    continue

                # Validate brand belongs to client
                brand = brands[brand_name]
                if brand.client_id != int(client_id):
                    errors.append(f"Row {row_idx}: Brand {brand_name} does not belong to Client ID {client_id}")
                    continue

                board = Board(
                    client=clients[client_id_str],
                    fiscal_year=fiscal_year,
                    address=address,
                    shop_name=shop_name,
                    brand=brands[brand_name],
                    board_type=board_types[board_type_name],
                    length=length,
                    breadth=breadth,
                    quantity=quantity,
                    rate=rate,
                    discount=discount
                )

                try:
                    board.full_clean()
                    boards_to_save.append(board)
                except ValidationError as e:
                    errors.append(f"Row {row_idx}: Validation error - {str(e)}")
                    continue

            if errors:
                return JsonResponse({'success': False, 'message': 'Errors in Excel data', 'errors': errors}, status=400)

            # Save all boards
            for board in boards_to_save:
                board.save()
                board.update_status()

            return JsonResponse({'success': True, 'message': f'Successfully uploaded {len(boards_to_save)} boards'})
    except Exception as e:
        logger.error(f"Error processing Excel upload: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error processing upload: {str(e)}'}, status=500)